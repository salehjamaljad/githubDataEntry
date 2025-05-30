import streamlit as st
import os
import pandas as pd
import pdfplumber
import zipfile
import tempfile
from io import BytesIO
from fuzzywuzzy import process
from datetime import datetime, timedelta
import pytz
import io
from docx import Document
from docx.shared import Inches
from docx.oxml.ns import qn
from docx.oxml import OxmlElement
from docx.shared import Inches, Pt
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl import Workbook
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, PatternFill, Side
from openpyxl.utils import get_column_letter
import os
import re
from streamlit_gsheets import GSheetsConnection
from config import translation_dict, categories_dict, branches_dict, branches_translation_tlbt, columns
def pdfToExcel():
    
    standardized_columns = [col.replace("\n", "_") for col in columns]
    conn = st.connection("gsheets", type=GSheetsConnection)
    

    # Special EG_ codes that need to capture the next word too
    special_codes = {
        "EG_Alex East_DS_", "EG_Alex", "EG_Zahraa Maadi", "EG_Nasrcity", "EG_Mansoura", 
        "EG_Tagamoa Golden", "EG_Tagamoa", "EG_Madinaty", "EG_Hadayek", "EG_October", "EG_Shrouk_", "EG_Mokatam", "EG_Sheikh"
    }

    def extract_eg_codes(pdf_path):
        with pdfplumber.open(pdf_path) as pdf:
            text = ""
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + " "  # Space to avoid word sticking issues

            # Tokenize the text by splitting on whitespace
            words = text.split()
            i = 0
            results = []

            while i < len(words):
                word = words[i]
                
                # Check if word starts with EG_ and is in special_codes
                if word.startswith("EG_"):
                    # If the word is in special_codes, capture the next word
                    if any(word == code or word.startswith(code) for code in special_codes):
                        next_word = words[i + 1] if i + 1 < len(words) else ""
                        combined = f"{word} {next_word}"
                        
                        # Fuzzy match for the closest branch in branches_dict
                        closest_match, score = process.extractOne(combined, branches_dict.keys())
                        if score >= 80:  # You can adjust the score threshold if needed
                            results.append({
                                "filename": os.path.basename(pdf_path), 
                                "extracted": combined, 
                                "matched_key": closest_match, 
                                "arabic_name": branches_dict[closest_match]
                            })
                        else:
                            results.append({"filename": os.path.basename(pdf_path), "extracted": combined})
                        
                        i += 1  # Skip the next word because it's already included
                    else:
                        # Fuzzy match for a single EG_ code if not in special_codes
                        closest_match, score = process.extractOne(word, branches_dict.keys())
                        if score >= 80:  # You can adjust the score threshold if needed
                            results.append({
                                "filename": os.path.basename(pdf_path), 
                                "extracted": word, 
                                "matched_key": closest_match, 
                                "arabic_name": branches_dict[closest_match]
                            })
                        else:
                            results.append({"filename": os.path.basename(pdf_path), "extracted": word})
                i += 1
            
            return results

    def process_pdf(file_path):
        all_tables = []
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    df = pd.DataFrame(table)
                    all_tables.append(df)

        for i, table in enumerate(all_tables):
            non_null_counts = table.notnull().sum()
            threshold = non_null_counts.max() * 0.5
            columns_to_drop = non_null_counts[non_null_counts <= threshold].index
            if len(columns_to_drop) > 0:
                all_tables[i] = table.drop(columns=columns_to_drop)

        for i, df in enumerate(all_tables[1:]):  # Skip first table if needed
            df.columns = standardized_columns

        if len(all_tables) > 1:
            final_df = pd.concat(all_tables[1:], ignore_index=True)
        else:
            final_df = all_tables[0]

        df = final_df
        df = df.loc[~(df.applymap(lambda x: x == '').all(axis=1))]
        df = df.reset_index(drop=True)
        df = df[df['Qty'] != '']
        df = df[df['SKU'] != 'SKU']
        # Drop the specified columns
        df.drop(columns=[
            'Disc._Amt.', 
            'Amt._Excl._VAT', 
            'VAT_%', 
            'VAT_Amt.', 
            'Supplier SKU',
            'No.',
            'Product'
        ], inplace=True)

        # Rename the specified columns
        df.rename(columns={
            'Unit_Cost': 'PP',
            'Amt._Incl._VAT': 'Total'
        }, inplace=True)

        # Convert data types
        df['PP'] = df['PP'].astype(float)
        df['Total'] = df['Total'].astype(float)
        df['Qty'] = df['Qty'].astype(int)
        try:
            df['Barcode'] = df['Barcode'].astype(int)
        except OverflowError:
            df['Barcode'] = df['Barcode'].astype(float)
        df['SKU'] = df['SKU'].astype(int)
        df["Item Name Ar"] = df["SKU"].map(translation_dict)
        df = df[['SKU', 'Barcode', 'Item Name Ar', 'PP', 'Qty', 'Total']]
        df = df.reset_index(drop=True)
        return df

    st.title("Purhcase Orders To Invoices")
    # Calculate the day after tomorrow
    default_date = datetime.today() + timedelta(days=2)

    # Use it as the default value
    selected_date = st.date_input('Enter the delivery date', value=default_date)
    
    df_invoice_number = conn.read(worksheet="Saved", cell="A1", ttl=5, headers=False)
    default_base_invoice_num = int(df_invoice_number.iat[0, 0])
    if "base_invoice_num" not in st.session_state:
        st.session_state.base_invoice_num = default_base_invoice_num
    base_invoice_num = st.number_input("رقم الفاتورة الأساسي", value=st.session_state.base_invoice_num, step=1)
    uploaded_zip = st.file_uploader("Upload a ZIP file containing PDFs", type=["zip"])

    if uploaded_zip is not None:
        with tempfile.TemporaryDirectory() as temp_dir:
            # Save the uploaded zip file
            zip_path = os.path.join(temp_dir, "uploaded.zip")
            with open(zip_path, "wb") as f:
                f.write(uploaded_zip.read())

            # Extract the zip
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                zip_ref.extractall(temp_dir)

            # Create a temp folder for Excel outputs
            output_dir = os.path.join(temp_dir, "excels")
            os.makedirs(output_dir, exist_ok=True)

            # Process each PDF
            pos_with_filenames = {}
            for filename in os.listdir(temp_dir):
                if filename.endswith(".pdf"):
                    file_path = os.path.join(temp_dir, filename)
                    df = process_pdf(file_path)
                    match = re.search(r"(PO\d+)", filename)
                    po = match.group(1) if match else None
                    pos_with_filenames[filename] = po

                    # Extract branch name for renaming
                    extracted_data = extract_eg_codes(file_path)
                    branch_name = None
                    if extracted_data:
                        branch_name = extracted_data[0].get("arabic_name", None)

                    # Use the branch name for renaming to Arabic only
                    if branch_name:
                        output_filename = f"{branch_name}_{po}_{selected_date}.xlsx"
                    else:
                        output_filename = f"{os.path.splitext(filename)[0]}.xlsx"

                    output_path = os.path.join(output_dir, output_filename)

                    # Save Excel without the 'po' column
                    df.to_excel(output_path, index=False, engine='openpyxl')

                    # Reopen and write PO in H1
                    wb = load_workbook(output_path)
                    ws = wb.active
                    ws["H1"] = po

                     # Clear 'Qty' and 'Total' columns if they exist (case-insensitive match)
                    for col in df.columns:
                        if col.strip().lower() == "qty":
                            df[col] = ''
                        if col.strip().lower() == "total":
                            df[col] = ''
                    # Add new sheet "فاتورة" and write df starting at A11
                    if "فاتورة" in wb.sheetnames:
                        del wb["فاتورة"]
                    ws_invoice = wb.create_sheet("فاتورة")
                    

                    # Define a thick border style
                    thick_border = Border(
                        left=Side(style='thick'),
                        right=Side(style='thick'),
                        top=Side(style='thick'),
                        bottom=Side(style='thick')
                    )
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )

                    # Write DataFrame to worksheet and apply thick border
                    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=11):
                        ws_invoice.row_dimensions[r_idx].height = 21  # Set row height to 35 pixels
                        for c_idx, value in enumerate(row, start=1):
                            cell = ws_invoice.cell(row=r_idx, column=c_idx, value=value)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            # If this is the Barcode column, apply number format to prevent scientific notation
                            if df.columns[c_idx - 1].lower() == 'barcode' and isinstance(value, (int, float)):
                                cell.number_format = '0'  # No decimals, no scientific notation


                    # Static cells
                    img = Image("Picture1.png")
                    ws_invoice.add_image(img, "A1")
                    ws_invoice["F1"] = "فاتورة مبيعات"
                    ws_invoice["F2"] = "رقم الفاتورة #"
                    ws_invoice["F3"] = "تاريخ الاستلام "
                    ws_invoice["E3"] = selected_date
                    ws_invoice["F4"] = "امر شراء رقم"
                    ws_invoice["E4"] = po
                    ws_invoice["F6"] = "اسم العميل "
                    ws_invoice["E6"] = "دليفيري هيرو ديمارت ايجيبت"
                    ws_invoice["F7"] = "الفرع"
                    ws_invoice["E7"] = branch_name
                    ws_invoice["C1"] = "شركه خضار للتجارة والتسويق"
                    ws_invoice["C1"].alignment = Alignment(horizontal='center', vertical='center')

                    ws_invoice["C2"] = "Khodar for Trading & Marketing"
                    ws_invoice["C2"].alignment = Alignment(horizontal='center', vertical='center')
                    ws_invoice["A5"] = "خضار.كوم"

                    # Apply thick borders to static cells
                    for cell_ref in ["F1", "F2", "F3", "E3", "F4", "E4", "F6", "E6", "F7", "E7", "E2", "C1", "C2", "A5"]:
                        ws_invoice[cell_ref].border = thick_border

                    # Add bottom cells and borders
                    df_end_row = 11 + len(df) + 1  # +1 for header row

                    center_align = Alignment(horizontal='center', vertical='center')

                    # Merge and write "Invoice Subtotal" in row df_end_row
                    ws_invoice.merge_cells(start_row=df_end_row, start_column=1, end_row=df_end_row, end_column=5)
                    ws_invoice.cell(row=df_end_row, column=1, value="Invoice Subtotal")
                    ws_invoice.cell(row=df_end_row, column=1).border = thick_border
                    ws_invoice.cell(row=df_end_row, column=1).alignment = center_align
                    ws_invoice.cell(row=df_end_row, column=6).border = thick_border

                    # Merge and write "Total" in row df_end_row + 1
                    ws_invoice.merge_cells(start_row=df_end_row + 1, start_column=1, end_row=df_end_row + 1, end_column=5)
                    ws_invoice.cell(row=df_end_row + 1, column=1, value="Total")
                    ws_invoice.cell(row=df_end_row + 1, column=1).border = thick_border
                    ws_invoice.cell(row=df_end_row + 1, column=1).alignment = center_align
                    ws_invoice.cell(row=df_end_row + 1, column=6).border = thick_border

                    ws_invoice.merge_cells(start_row=df_end_row + 3, start_column=1, end_row=df_end_row + 3, end_column=6)
                    ws_invoice.cell(row=df_end_row + 3, column=1, value="شركة خضار للتجارة و التسويق ")
                    ws_invoice.cell(row=df_end_row + 3, column=1).alignment = center_align

                    ws_invoice.merge_cells(start_row=df_end_row + 4, start_column=1, end_row=df_end_row + 4, end_column=6)
                    ws_invoice.cell(row=df_end_row + 4, column=1, value="        ش.ذ.م.م")
                    ws_invoice.cell(row=df_end_row + 4, column=1).alignment = center_align

                    ws_invoice.merge_cells(start_row=df_end_row + 5, start_column=1, end_row=df_end_row + 5, end_column=6)
                    ws_invoice.cell(row=df_end_row + 5, column=1, value="سجل تجارى / 13138  بطاقه ضريبية/721/294/448")
                    ws_invoice.cell(row=df_end_row + 5, column=1).alignment = center_align


                    # Adjust column widths based on max length of content in each column
                    for col in ws_invoice.columns:
                        max_length = 0
                        column = col[0].column  # Get the column number
                        column_letter = get_column_letter(column)
                        
                        if column_letter == 'A':  # Check if the column is "A"
                            ws_invoice.column_dimensions[column_letter].width = 10  # Set column A width to 10
                        else:
                            for cell in col:
                                try:
                                    if cell.value:
                                        max_length = max(max_length, len(str(cell.value)))
                                except:
                                    pass
                            adjusted_width = max_length + 2  # Add padding
                            ws_invoice.column_dimensions[column_letter].width = adjusted_width

                    # Loop through all used cells and apply bold formatting
                    for row in ws_invoice.iter_rows():
                        for cell in row:
                            if cell.value or cell.coordinate == "E2":
                                cell.font = Font(bold=True)

                    # Save workbook
                    wb.save(output_path)


            
            all_dfs = []

            for excel_file in os.listdir(output_dir):
                if excel_file.endswith(".xlsx"):
                    excel_path = os.path.join(output_dir, excel_file)
                    df = pd.read_excel(excel_path, usecols=range(6))  # Read only first 6 columns

                    # Get branch name and PO from file name
                    base = os.path.splitext(excel_file)[0]
                    parts = base.split("_")
                    if len(parts) >= 2:
                        branch_name = parts[0]
                        po = parts[1]
                        df["branch"] = branch_name
                        df["po"] = po
                    

                    all_dfs.append(df)

            if all_dfs:
                combined_df = pd.concat(all_dfs, ignore_index=True)
                # Ensure SKU is integer
                combined_df["SKU"] = pd.to_numeric(combined_df["SKU"], errors="coerce").astype("Int64")

                # Replace Product column using SKU mapped through translation_dict
                combined_df["Product"] = combined_df["SKU"].map(translation_dict)
                # Create a reverse mapping: product -> category
                reverse_categories = {
                    item: category for category, items in categories_dict.items() for item in items
                }

                # Map the Product column to its category
                combined_df["category"] = combined_df["Product"].map(reverse_categories)

                st.dataframe(combined_df)



                # Pivot the dataframe
                pivot_df = combined_df.pivot_table(
                    index=["Barcode", "SKU", "Product", "category", "PP"],
                    columns="branch",
                    values="Qty",
                    aggfunc="sum",
                    fill_value=0
                ).reset_index()

                # Rename 'Product' to 'Product name' for consistency
                pivot_df = pivot_df.rename(columns={"Product": "Product name"})
                pivot_df[sorted(pivot_df.columns)]
                # Define column groups
                alexandria_columns = ["Barcode",'Product name', 'SKU', 'category', "PP",'سيدي بشر', 'الابراهيميه', 'وينجت']
                ready_veg_columns = ["Barcode",'Product name', 'SKU', 'category',  "PP", 'المعادي لاسلكي', 'الدقي', 'زهراء المعادي',
                                    'ميدان لبنان', 'العجوزة', 'كورنيش المعادي', 'زهراء المعادي - 2', "الظاهر", "المقطم", "السيدة زينب", "حلوان",
                                    "المنيل", "المقطم 2 هضبة", "شبرا", "زايد 2", "حدائق الاهرام", "اكتوبر",
                                    "الشيخ زايد", "بالم هيلز", "سيتي ستارز", "هيليوبليس"]
                
                
                # Always-include base columns
                base_columns = ["Barcode", 'Product name', 'SKU', 'category', "PP"]

                # Get all unique branch columns used in Alex and Ready Veg
                used_branch_columns = set(alexandria_columns + ready_veg_columns) - set(base_columns)

                # All other branch columns are considered Cairo
                cairo_branch_columns = [col for col in pivot_df.columns if col not in used_branch_columns and col not in base_columns]

                # Now construct the final Cairo columns list
                cairo_columns = base_columns + cairo_branch_columns


                # Create the split DataFrames
                alexandria_df = pivot_df[[col for col in alexandria_columns if col in pivot_df.columns]]
                ready_veg_df = pivot_df[[col for col in ready_veg_columns if col in pivot_df.columns]]
                cairo_df = pivot_df[[col for col in cairo_columns if col in pivot_df.columns]]
                def reorder_columns(df):
                    # Columns to always appear first
                    first_cols = ['Barcode', 'SKU', 'Product name']
                    
                    # Columns to appear last
                    last_cols = ['PP', 'category']
                    
                    # Remaining columns (excluding first and last), sorted alphabetically
                    middle_cols = sorted([col for col in df.columns if col not in first_cols + last_cols])
                    
                    # Final ordered list
                    ordered_cols = first_cols + middle_cols + last_cols
                    return df[[col for col in ordered_cols if col in df.columns]]  # filter to existing columns


                # Apply to each dataframe
                alexandria_df = reorder_columns(alexandria_df)
                ready_veg_df = reorder_columns(ready_veg_df)
                cairo_df = reorder_columns(cairo_df)



                # Define category sort order
                category_order = {
                    "فاكهه": 1,
                    "خضار": 2,
                    "جاهز": 3,
                    "اعشاب": 4
                }

                def add_total_and_sort(df):
                    # Identify branch columns by excluding fixed columns
                    fixed_cols = ["Barcode", 'Product name', 'SKU', 'category', "PP"]
                    branch_cols = [col for col in df.columns if col not in fixed_cols]

                    # Calculate total quantity and total value
                    df["total quantity"] = df[branch_cols].sum(axis=1)
                    df["total"] = df["PP"] * df["total quantity"]

                    # Map sort key for category
                    df["category_order"] = df["category"].map(category_order)

                    # Sort by category order and product name
                    df = df.sort_values(by=["category_order", "Product name"], ascending=[True, True])

                    # Drop helper column
                    df = df.drop(columns=["category_order"])

                    # Reorder columns: insert total quantity before PP, and total after PP
                    cols = df.columns.tolist()
                    try:
                        pp_index = cols.index("PP")
                    except ValueError:
                        pp_index = 0  # Fallback

                    # Remove total and total quantity from their original position
                    cols.remove("total quantity")
                    cols.remove("total")

                    # Insert total quantity before PP, and total after PP
                    cols = cols[:pp_index] + ["total quantity", "PP", "total"] + cols[pp_index+1:]

                    # Reorder the DataFrame
                    df = df[cols]

                    return df


                # Apply to each DataFrame
                alexandria_df = add_total_and_sort(alexandria_df)
                ready_veg_df = add_total_and_sort(ready_veg_df)
                cairo_df = add_total_and_sort(cairo_df)
                def append_grand_total(df):
                    if not {"total quantity", "PP", "total"}.issubset(df.columns):
                        return df

                    cols = df.columns.tolist()

                    # Get indexes
                    try:
                        product_name_idx = cols.index("Product name")
                        pp_idx = cols.index("PP")
                    except ValueError:
                        return df  # Required columns missing

                    # Columns to sum: between 'Product name' and 'PP' (exclusive)
                    sum_columns = cols[product_name_idx + 1:pp_idx]

                    grand_total_row = {col: "" for col in df.columns}
                    grand_total_row["Product name"] = "Grand Total"

                    # Sum intermediate columns
                    for col in sum_columns:
                        if pd.api.types.is_numeric_dtype(df[col]):
                            grand_total_row[col] = df[col].sum()

                    # Sum fixed known columns
                    grand_total_row["total quantity"] = df["total quantity"].sum()
                    grand_total_row["PP"] = df["PP"].sum()
                    grand_total_row["total"] = df["total"].sum()

                    df = pd.concat([df, pd.DataFrame([grand_total_row])], ignore_index=True)
                    return df

                # Filter out zero total rows
                alexandria_df = alexandria_df[alexandria_df["total"] != 0]
                ready_veg_df = ready_veg_df[ready_veg_df["total"] != 0]
                cairo_df = cairo_df[cairo_df["total"] != 0]

                # Append grand total row
                alexandria_df = append_grand_total(alexandria_df)
                ready_veg_df = append_grand_total(ready_veg_df)
                cairo_df = append_grand_total(cairo_df)



                
                
                
                
                
                
                
                # Add Cairo DF
                cairo_buffer = BytesIO()
                with pd.ExcelWriter(cairo_buffer, engine='xlsxwriter') as writer:
                    cairo_df.to_excel(writer, index=False)
                
                # Add Ready Veg DF
                ready_buffer = BytesIO()
                with pd.ExcelWriter(ready_buffer, engine='xlsxwriter') as writer:
                    ready_veg_df.to_excel(writer, index=False)
                
                # Add Alexandria DF
                alex_buffer = BytesIO()
                with pd.ExcelWriter(alex_buffer, engine='xlsxwriter') as writer:
                    alexandria_df.to_excel(writer, index=False)
                
                # Create ZIP
                output_zip_buffer = BytesIO()
                with zipfile.ZipFile(output_zip_buffer, "w") as zipf:
                    # Add Excel files from output_dir
                    for excel_file in os.listdir(output_dir):
                        excel_path = os.path.join(output_dir, excel_file)
                        zipf.write(excel_path, arcname=excel_file)

                    # Add in-memory Excel dataframes
                    zipf.writestr(f"مجمع_طلبات_اسكندرية_{selected_date}.xlsx", alex_buffer.getvalue())
                    zipf.writestr(f"مجمع_طلبات_الخضار_الجاهز_{selected_date}.xlsx", ready_buffer.getvalue())
                    zipf.writestr(f"مجمع_طلبات_القاهرة_{selected_date}.xlsx", cairo_buffer.getvalue())

                # After all processing (creating Excel files and updating branch_offsets)
                # Now assign invoice numbers based on branch_offsets and update E2 in the Excel files

                special_branches = ["الابراهيميه", "سيدي بشر", "وينجت"]
                branch_offsets = {}

                # Step 1: Gather all .xlsx files
                filenames = [f for f in os.listdir(output_dir) if f.endswith(".xlsx")]

                # Step 2: Map filenames to branch names
                file_branch_map = {filename: filename.split("_")[0] for filename in filenames}

                # Step 3: Determine the ordered list of branches
                present_specials = [b for b in special_branches if b in file_branch_map.values()]
                other_branches = sorted(set(file_branch_map.values()) - set(special_branches))

                # Step 4: Assign offsets based on required priority
                offset = 0
                for b in present_specials + other_branches:
                    branch_offsets[b] = offset
                    offset += 1

                # Step 5: Assign invoice number and update Excel files
                for filename, branch_name in file_branch_map.items():
                    final_invoice_number = base_invoice_num + branch_offsets.get(branch_name, 0)
                    output_path = os.path.join(output_dir, filename)

                    wb = load_workbook(output_path)
                    if "فاتورة" in wb.sheetnames:
                        ws = wb["فاتورة"]
                        ws["E2"] = final_invoice_number
                        wb.save(output_path)

               # Create a new workbook for consolidation
                consolidated_wb = Workbook()
                consolidated_wb.remove(consolidated_wb.active)

                for filename in filenames:
                    file_path = os.path.join(output_dir, filename)
                    wb = load_workbook(file_path, data_only=True)
                    if "فاتورة" in wb.sheetnames:
                        source_ws = wb["فاتورة"]
                        new_sheet_name = os.path.splitext(filename)[0][:31]  # Sheet name max length is 31
                        target_ws = consolidated_wb.create_sheet(title=new_sheet_name)
                        for merged_range in source_ws.merged_cells.ranges:
                            target_ws.merge_cells(str(merged_range))

                        for row in source_ws.iter_rows():
                            for cell in row:
                                new_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)

                                # Copy styles safely
                                new_cell.font = Font(
                                    name=cell.font.name,
                                    size=cell.font.size,
                                    bold=cell.font.bold,
                                    italic=cell.font.italic,
                                    vertAlign=cell.font.vertAlign,
                                    underline=cell.font.underline,
                                    strike=cell.font.strike,
                                    color=cell.font.color
                                )
                                new_cell.alignment = Alignment(
                                    horizontal=cell.alignment.horizontal,
                                    vertical=cell.alignment.vertical,
                                    wrap_text=cell.alignment.wrap_text
                                )
                                new_cell.border = Border(
                                    left=cell.border.left,
                                    right=cell.border.right,
                                    top=cell.border.top,
                                    bottom=cell.border.bottom
                                )
                                new_cell.fill = PatternFill(
                                    fill_type=cell.fill.fill_type,
                                    fgColor=cell.fill.fgColor,
                                    bgColor=cell.fill.bgColor
                                )
                                new_cell.number_format = cell.number_format
                        # Step 3: Copy row heights
                        for row in source_ws.iter_rows():
                            target_ws.row_dimensions[row[0].row].height = source_ws.row_dimensions[row[0].row].height

                        # Step 4: Copy column widths
                        for col in source_ws.columns:
                            column_letter = get_column_letter(col[0].column)
                            target_ws.column_dimensions[column_letter].width = source_ws.column_dimensions[column_letter].width
                        # Step 5: Add picture to A1 in the new sheet
                        img = Image("Picture1.png")
                        target_ws.add_image(img, 'A1')
               

                # Save the consolidated workbook to a BytesIO object
                invoices_buffer = BytesIO()
                consolidated_wb.save(invoices_buffer)
                invoices_buffer.seek(0)

                # Collect totals for PO summary
                po_summary = []

                # Names of generated buffers to exclude from scanning
                excluded_files = {
                    f"مجمع_طلبات_اسكندرية_{selected_date}.xlsx",
                    f"مجمع_طلبات_الخضار_الجاهز_{selected_date}.xlsx",
                    f"مجمع_طلبات_القاهرة_{selected_date}.xlsx",
                    "فواتير.xlsx"
                }

                for filename in os.listdir(output_dir):
                    if filename in excluded_files or not filename.endswith(".xlsx"):
                        continue

                    file_path = os.path.join(output_dir, filename)
                    wb = load_workbook(file_path, data_only=True)

                    if "Sheet1" not in wb.sheetnames:
                        continue

                    ws = wb["Sheet1"]
                    h1_text = ws["H1"].value

                    # Find the "Total" column in the header row
                    total_col_letter = None
                    total_col_idx = None  # will store the column index (1-based)
                    for cell in ws[1]:
                        if cell.value and str(cell.value).strip().lower() == "total":
                            total_col_letter = get_column_letter(cell.column)
                            total_col_idx = cell.column
                            break

                    if total_col_letter is None:
                        continue

                    total_sum = 0
                    for row in ws.iter_rows(min_row=2):
                        val = row[total_col_idx - 1].value
                        if isinstance(val, (int, float)):
                            total_sum += val

                    # Get invoice number from second sheet "فاتورة"
                    invoice_number = None
                    if "فاتورة" in wb.sheetnames:
                        invoice_ws = wb["فاتورة"]
                        invoice_val = invoice_ws["E2"].value
                        if isinstance(invoice_val, int):
                            invoice_number = invoice_val

                    # Extract Arabic branch from filename (assuming it's the part before "_")
                    arabic_branch = filename.split("_")[0]

                    # Get English branch name using dict, fallback to arabic_branch if not found
                    english_branch = branches_translation_tlbt.get(arabic_branch, arabic_branch)

                    po_summary.append((english_branch, arabic_branch,  h1_text, total_sum, invoice_number))


                # Create po_totals.xlsx in memory
                po_totals_wb = Workbook()
                po_ws = po_totals_wb.active
                po_ws.title = "Summary"

                # Add header with new "branch name (en)" column at the beginning
                po_ws.append(["branch (en)", "branch (ar)", "po", "Total of the po", "invoice_number"])

                for item in po_summary:
                    po_ws.append(item)

                po_totals_buffer = BytesIO()
                po_totals_wb.save(po_totals_buffer)
                po_totals_buffer.seek(0)
                # After all Excel files have been updated, create the ZIP file with the updated Excel files
                output_zip_buffer = BytesIO()
                with zipfile.ZipFile(output_zip_buffer, "w") as zipf:

                    # Create inner zip buffer
                    inner_zip_buffer = BytesIO()
                    with zipfile.ZipFile(inner_zip_buffer, "w") as inner_zip:
                        for excel_file in os.listdir(output_dir):
                            if excel_file not in excluded_files and excel_file.endswith(".xlsx"):
                                excel_path = os.path.join(output_dir, excel_file)
                                inner_zip.write(excel_path, arcname=excel_file)

                    inner_zip_buffer.seek(0)
                    zipf.writestr(f"ملفات الفروع_{selected_date}.zip", inner_zip_buffer.getvalue())

                    # Add the excluded files to the main ZIP
                    zipf.writestr(f"po_totals_{selected_date}.xlsx", po_totals_buffer.getvalue())
                    zipf.writestr(f"مجمع_طلبات_اسكندرية_{selected_date}.xlsx", alex_buffer.getvalue())
                    zipf.writestr(f"مجمع_طلبات_الخضار_الجاهز_{selected_date}.xlsx", ready_buffer.getvalue())
                    zipf.writestr(f"مجمع_طلبات_القاهرة_{selected_date}.xlsx", cairo_buffer.getvalue())
                    zipf.writestr("فواتير.xlsx", invoices_buffer.getvalue())

                output_zip_buffer.seek(0)
                excluded_files = {
                    f"po_totals_{selected_date}.xlsx",
                    f"مجمع_طلبات_اسكندرية_{selected_date}.xlsx",
                    f"مجمع_طلبات_الخضار_الجاهز_{selected_date}.xlsx",
                    f"مجمع_طلبات_القاهرة_{selected_date}.xlsx",
                    "فواتير.xlsx"
                }


                st.success("Processing complete!")
                st.info(f"last invoice number generated: {offset + base_invoice_num-1}")
                
                if st.download_button(
                    label="Download All Files as ZIP",
                    data=output_zip_buffer.getvalue(),
                    file_name=f"talabat_documents_{selected_date}.zip",
                    mime="application/zip"
                ):
                    if base_invoice_num == default_base_invoice_num:
                        df_invoice_number.iat[0, 0] = offset + int(df_invoice_number.iat[0, 0])
                    else:
                        df_invoice_number.iat[0, 0] = default_base_invoice_num
                    conn.update(worksheet="Saved", data=df_invoice_number)

if __name__ == "__main__":
    pdfToExcel()
